'''This file handles all data writing'''


import datetime
import json
import matplotlib.pyplot as plt
import numpy as np
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter
import os
import pandas as pd
import random
from rich import print as rprint
import xlsxwriter


def generate_report(data, file_path):
    '''A function that appends results to file
    '''
    

    current_time = datetime.datetime.now() # get datestamp
    timestamp = current_time.strftime('%Y-%m-%d')

    # Extract unique models, constants, thicknesses (preserving order)
    model_constants = {}
    thicknesses = set()
    temperatures = list(data.keys())

    for temp_data in data.values():
        for thick, thickness_data in temp_data.items():
            thicknesses.add(thick)
            for model, model_data in thickness_data.items():
                if model not in model_constants:
                    model_constants[model] = []
                for const in model_data["Constants"].keys():
                    if const not in model_constants[model]:
                        model_constants[model].append(const)

    # Convert thicknesses to a list (sorted if needed)
    thicknesses = sorted(list(thicknesses))

    # Prepare rows and track model boundaries
    rows = []
    sn = 1
    model_boundaries = {}  # {model: (first_row, last_row)}

    current_row = 3  # Starting after headers
    for model in model_constants:
        constants = model_constants[model]
        first_row = current_row
        for i, const in enumerate(constants):
            row = [sn if i == 0 else '', model if i == 0 else '', const]
            for temp in temperatures:
                for thick in thicknesses:
                    value = ""
                    try:
                        value = data[temp][thick][model]["Constants"].get(const, "")
                    except KeyError:
                        pass
                    row.append(value)
            rows.append(row)
            current_row += 1
        model_boundaries[model] = (first_row, current_row - 1)
        sn += 1

    # Create workbook and sheet
    # wb = Workbook()
    # ws = wb.active
     # check if file exists
    if os.path.exists(file_path):
        wb = load_workbook(file_path)
    else:
        wb = Workbook()

        # Remove default sheet if it's empty
        if 'Sheet' in workbook.sheetnames:
            wb.remove(workbook['Sheet'])

    # Add new sheet
    ws = wb.create_sheet(title='Model Constants')

    # Define styles
    center = Alignment(horizontal="center", vertical="center")
    bold_font = Font(bold=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    thick_bottom_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thick')
    )

    # Create header rows
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
    ws.merge_cells(start_row=1, start_column=2, end_row=2, end_column=2)
    ws.merge_cells(start_row=1, start_column=3, end_row=2, end_column=3)

    ws.cell(row=1, column=1).value = "S/N"
    ws.cell(row=1, column=2).value = "Models"
    ws.cell(row=1, column=3).value = "Constants"

    ws.cell(row=1, column=1).font = bold_font
    ws.cell(row=1, column=2).font = bold_font
    ws.cell(row=1, column=3).font = bold_font

    col = 4
    for temp in temperatures:
        ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + len(thicknesses) - 1)
        ws.cell(row=1, column=col).value = f'{temp} °C'
        ws.cell(row=1, column=col).font = bold_font
        for i, thick in enumerate(thicknesses):
            cell = ws.cell(row=2, column=col + i)
            cell.value = f'{thick} mm'
            cell.font = bold_font
            cell.alignment = center
            cell.border = thin_border
        col += len(thicknesses)

    # Align and border header cells
    for row in ws.iter_rows(min_row=1, max_row=2):
        for cell in row:
            cell.alignment = center
            cell.border = thin_border

    # Add the data with proper borders
    for r_idx, row_data in enumerate(rows, start=3):
        is_model_last_row = any(r_idx == last for (_, last) in model_boundaries.values())
        
        for c_idx, val in enumerate(row_data, start=1):
            cell = ws.cell(row=r_idx, column=c_idx)
            cell.value = val
            cell.alignment = center
            
            # Apply appropriate border
            if is_model_last_row:
                cell.border = thick_bottom_border
            else:
                cell.border = thin_border

            # Bold model names
            if c_idx == 2 and val != '':
                cell.font = bold_font


    # filename = f'{folder_path}/model_constants_{timestamp}.xlsx'

    try:
        # Save file
        wb.save(file_path)
    except Exception as e:
        print(e)
        print('Error found when writing model constants')
    else:
        print(f'Data has been written to {file_path}')

    return 0


def Deff_writer(data, temperatures, thicknesses, folder_path):
    '''
    This function writes moisture diffusivity data
    to an excel sheet
    '''
    current_time = datetime.datetime.now() # get datestamp
    timestamp = current_time.strftime('%Y-%m-%d')

    # append datestamp to file
    file_name = f'{folder_path}/Analysis_result_{timestamp}.xlsx'
    
    # === Workbook Setup ===
    wb = Workbook()
    ws = wb.active
    ws.title = "Moisture Diffusivity"

    # === Styling ===
    bold_center = Font(bold=True)
    center = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # === Title Row (Row 1) ===
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=1 + len(temperatures))
    ws.cell(row=1, column=1).value = "Moisture Diffusivity Data"
    ws.cell(row=1, column=1).font = Font(bold=True, size=14)
    ws.cell(row=1, column=1).alignment = center

    # === Header Row 1 (Row 2) ===
    ws.cell(row=2, column=1).value = "Thickness (mm)"
    ws.cell(row=2, column=1).font = bold_center
    ws.cell(row=2, column=1).alignment = center
    ws.merge_cells(start_row=2, start_column=2, end_row=2, end_column=1 + len(temperatures))
    ws.cell(row=2, column=2).value = "Moisture diffusivity of samples (m^2/s)"
    ws.cell(row=2, column=2).alignment = center
    ws.cell(row=2, column=2).font = bold_center

    # === Header Row 2 (Row 3) ===
    for col_index, temp in enumerate(temperatures, start=2):
        ws.cell(row=3, column=col_index).value = f"{temp} °C"
        ws.cell(row=3, column=col_index).alignment = center
        ws.cell(row=3, column=col_index).font = bold_center

    # === Data Rows (Row 4+) ===
    for row_index, thickness in enumerate(thicknesses, start=4):
        ws.cell(row=row_index, column=1).value = thickness
        ws.cell(row=row_index, column=1).alignment = center
        for col_index, value in enumerate(data[row_index - 4], start=2):
            ws.cell(row=row_index, column=col_index).value = value
            ws.cell(row=row_index, column=col_index).alignment = center


    # === Apply Borders ===
    max_row = 3 + len(thicknesses)
    max_col = 1 + len(temperatures)
    for row in ws.iter_rows(min_row=2, max_row=max_row, min_col=1, max_col=max_col):
        for cell in row:
            cell.border = thin_border

    # === Adjust Column Widths ===
    for col_letter in ['A', 'B', 'C', 'D'][:max_col]:
        ws.column_dimensions[col_letter].width = 18


    try:
        # === Save Workbook ===
        wb.save(file_name)
    except Exception as e:
        print('Error writing Moisture Diffusivity to file')
    else:
        print(f'Moisture Diffusivity Data has been written successfully to {file_name}')

    return file_name



def gen_act_energy_report(Ea_data, thicknesses, file_path):
    '''
    This function writes the activation
    energy result to an excel sheet
    '''
    # check if file exists
    if os.path.exists(file_path):
        wb = load_workbook(file_path)
    else:
        wb = Workbook()

        # Remove default sheet if it's empty
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])

    # Add new sheet
    ws = wb.create_sheet(title=f'Thermodynamics result')
  
    # styling
    bold = Font(bold=True)
    title_font = Font(bold=True, size=14)
    center = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    num_cols = len(thicknesses) + 1  # +1 for the first column

     # === Title Row ===
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
    ws.cell(row=1, column=1).value = "Thermodynamics Data"
    ws.cell(row=1, column=1).font = title_font
    ws.cell(row=1, column=1).alignment = center

    # === Header Row (Thickness) ===
    ws.cell(row=2, column=1).value = "Thickness (mm)"
    ws.cell(row=2, column=1).font = bold
    ws.cell(row=2, column=1).alignment = center
    for col, thickness in enumerate(thicknesses, start=2):
        cell = ws.cell(row=2, column=col)
        cell.value = thickness
        cell.font = bold
        cell.alignment = center

    # === Data Row (Ea values) ===
    ws.cell(row=3, column=1).value = "Activation energy Ea (kJ/mol)"
    ws.cell(row=3, column=1).font = bold
    ws.cell(row=3, column=1).alignment = center
    for col, value in enumerate(Ea_data, start=2):
        cell = ws.cell(row=3, column=col)
        cell.value = value
        cell.alignment = center
   
    # === Apply borders ===
    for row in ws.iter_rows(min_row=2, max_row=3, min_col=1, max_col=num_cols):
        for cell in row:
            cell.border = thin_border

    # === Adjust column widths ===
    for col_index in range(1, num_cols + 1):
        col_letter = chr(64 + col_index)  # e.g., 1 → 'A', 2 → 'B'
        ws.column_dimensions[col_letter].width = 18


    try:
        # === Save Workbook ===
        wb.save(file_path)
    except Exception as e:
        print('Error writing Activation Energy to file')
    else:
        print(f'Activation Energy Data has been written successfully to {file_path}')

    return 0


def custom_csv_writer(temperatures, thicknesses, data, header, file_path):
    '''
    this function writes data to a csv file
    '''
    try:
        # Load workbook and sheet
        wb = load_workbook(file_path)
        sheet = wb['Thermodynamics result']
        last_row = sheet.max_row
        
        # Style definitions
        bold_font = Font(bold=True)
        center_aligned = Alignment(horizontal='center')
        
        # Define border styles
        full_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )    
  
        # Create header structure (2 rows)
        
        # Row 1: Empty cell + Merged "Enthalpy (j/mol)" header
        sheet.cell(row=last_row+2, column=1).border = full_border
        
        # Merge and format enthalpy header
        sheet.merge_cells(
            start_row=last_row+2,
            start_column=2,
            end_row=last_row+2,
            end_column=len(temperatures)+1
        )
        header_cell = sheet.cell(
            row=last_row+2, 
            column=2, 
            value=header
        )

        header_cell.font = bold_font
        header_cell.alignment = center_aligned
        header_cell.border = full_border  # Apply full border to merged cell
        
        # Apply borders to all cells in the merged range
        for col in range(2, len(temperatures)+2):
            cell = sheet.cell(row=last_row+2, column=col)
            cell.border = full_border

         # Row 2: "Thickness" + Temperature headers
        sheet.cell(
            row=last_row + len(temperatures), 
            column=1, 
            value="Thickness"
        ).font = bold_font
        
        for col_idx, temp in enumerate(temperatures, start=2):
            temp_cell = sheet.cell(
                row=last_row+3,
                column=col_idx,
                value=f"{temp} °C"
            )
            temp_cell.font = bold_font
            temp_cell.alignment = center_aligned
            temp_cell.border = full_border
        
        # Data rows
        for row_idx, (thickness, row_data) in enumerate(zip(thicknesses, data), start=last_row+4):
            # Thickness column
            sheet.cell(
                row=row_idx,
                column=1,
                value=f"{thickness}mm"
            ).border = full_border
            
            # Data columns
            for col_idx, value in enumerate(row_data, start=2):
                sheet.cell(
                    row=row_idx,
                    column=col_idx,
                    value=value
                ).border = full_border
        
        # Column formatting
        sheet.column_dimensions['A'].width = 10  # Thickness column
        for col in range(2, len(temperatures)+2):
            sheet.column_dimensions[get_column_letter(col)].width = 12
        
        # Save changes
        wb.save(file_path)
        print(f"{header} appended successfully to {file_path}")
        
    except Exception as e:
        print(f"Error: {str(e)}")

    return 0
        

def model_report_writer(data, temp):
    '''
    pass
    '''
    result = []
    # main_key = list(data.keys())[temp] # get first level keys (temperature)
    thickness_list = list(data[temp].keys()) # get second level keys (thickness)

    models = list(data[temp][thickness_list[0]].keys()) # get models

    for i in range(len(models)):

        _ = [i+1,models[i]]

        for thickness in thickness_list:
            # get statistical indicators
            r_square = data[temp][thickness][models[i]]['R_Square']
            sse = data[temp][thickness][models[i]]['SSE']
            rmse = data[temp][thickness][models[i]]['RMSE']


            _.append(r_square)
            _.append(sse)
            _.append(rmse)

        result.append(_)

    return result


def create_dynamic_table(filename, file_path, main_headers, sub_headers, data, temp, thickness_list, best_model_list):
    '''
    creates and formats the models report sheet
    '''
    # create title
    title = f'Result summary of the statistical curve fitting analysis at {temp} °C'
    
    # check if file exists
    if os.path.exists(file_path):
        wb = load_workbook(file_path)
    else:
        wb = Workbook()

        # Remove default sheet if it's empty
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])

    # Add new sheet
    ws = wb.create_sheet(title=f'Model Performance@{temp} °C')

    # merge 5 columns in the first row
    ws.merge_cells('A1:H1')

    # input title in merged cell
    ws['A1'] = title

    # Align title
    # ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws['A1'].font = Font(bold=True)

    # Dynamically set the main headers and merge cells
    col_start = 3  # Start column for main headers
    for i, header in enumerate(main_headers):
        col_end = col_start + len(sub_headers) - 1
        ws.merge_cells(start_row=2, start_column=col_start, end_row=2, end_column=col_end)
        ws.cell(row=2, column=col_start).value = header
        ws.cell(row=2, column=col_start).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=2, column=col_start).font = Font(bold=True)
        col_start = col_end + 1

    # Add the sub-headers below the main headers
    ws.append(["No.", "Model Name"] + sub_headers * len(main_headers))

    # Apply bold font and center alignment to sub-headers
    for cell in ws[3]:
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.font = Font(bold=True)

    # Append the data rows
    for row in data:
        ws.append(row)

    # Apply center alignment and borders to the entire table
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    for row in ws.iter_rows(min_row=2, max_row=len(data) + 3 , min_col=1, max_col=2 + len(main_headers) * len(sub_headers)):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

    

    # Handle writing of best model details
    
    # Styling
    thin = Side(border_style="thin", color="000000")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)
    bold_font = Font(bold=True)
    center_alignment = Alignment(horizontal="center", vertical="center")

    row = 25

    # Pair each thickness with its corresponding model
    for thickness, model in zip(thickness_list, best_model_list):
        # Write header row
        ws.cell(row=row, column=1, value="Model Name").font = bold_font
        ws.cell(row=row, column=2, value="constants").font = bold_font
        ws.cell(row=row, column=3, value=f"{thickness}mm").font = bold_font
        
        # Apply styling to header
        for col in range(1, 4):
            ws.cell(row=row, column=col).border = border
            ws.cell(row=row, column=col).alignment = center_alignment
            ws.column_dimensions[get_column_letter(col)].width = 15
        
        row += 1

        # Write model name row (first constant)
        constants = list(model['constants'].items())
        ws.cell(row=row, column=1, value=model['model name']).border = border
        ws.cell(row=row, column=2, value=constants[0][0]).border = border  # 'a'
        ws.cell(row=row, column=3, value=constants[0][1]).border = border  # value
        row += 1

        # Write model eqn row (second constant)
        ws.cell(row=row, column=1, value=model["model equation"]).border = border
        ws.cell(row=row, column=2, value=constants[1][0]).border = border  # 'b'
        ws.cell(row=row, column=3, value=constants[1][1]).border = border  # value
        row += 1

        # Write remaining constants if they exist
        for const_name, const_value in constants[2:]:
            ws.cell(row=row, column=1, value="").border = border
            ws.cell(row=row, column=2, value=const_name).border = border
            ws.cell(row=row, column=3, value=const_value).border = border
            row += 1

        # Add blank row between sections
        row += 1


    # Save the workbook
    wb.save(file_path)

    return file_path


# def df_writer(df_list, folder_path):
#     '''
#     function takes a list of dataframes and saves
#     them to different sheets in one excel file.
#     '''
#     # create timestamp
#     current_time = datetime.datetime.now()
#     timestamp = current_time.strftime('%Y-%m-%d')
#     file_name = 'Analysis Results'
#     file_path = f'{folder_path}/{file_name}_{timestamp}.xlsx'
#     df_list[0].to_excel(file_path)
   
    
#     # Create an Excel writer object
#     with pd.ExcelWriter(file_path, engine="xlsxwriter") as writer:
#         df_list[0].to_excel(writer, sheet_name="Moisture Diffusivity",)  # First sheet

#         if all(i is not None for i in df_list[1:]):
#             df_list[1].to_excel(writer, sheet_name="Activation Energy", index=False)  # Second sheet
#             df_list[2].to_excel(writer, sheet_name="Enthalpy" )  # Third sheet
#             df_list[3].to_excel(writer, sheet_name="Entropy")  # Forth sheet
#             df_list[4].to_excel(writer, sheet_name="Gibbs Free Energy")  # Fifth sheet

    
#     print("Excel file with multiple sheets saved successfully!")

#     return file_path


def df_writer(df_list, folder_path):
    '''
    function takes a list of dataframes and saves
    them to different sheets in one excel file.
    '''
    # create timestamp
    current_time = datetime.datetime.now()
    timestamp = current_time.strftime('%Y-%m-%d')
    file_name = f'{folder_path}/Thermodynamics_Results_{timestamp}.csv'


    with open(file_name, "w", newline='') as f:
        f.write("Moisture Diffusivity Data\n")  # Add a header
        df_list[0].to_csv(f)

        if all(i is not None for i in df_list[1:]):
            f.write("\nActivation Energy Data\n")
            df_list[1].to_csv(f, index=False)

            f.write("\nEnthalpy Data\n")
            df_list[2].to_csv(f)

            f.write("\nEntropy Data\n")
            df_list[3].to_csv(f)

            f.write("\nGibbs Free Energy Data\n")
            df_list[4].to_csv(f)

    
    print("Excel file with multiple sheets saved successfully!")

    return f'{folder_path}/results.xlsx'



def make_dir():
    '''
    This function creates a result folder to 
    contain all results generated
    '''
    from pathlib import Path
    import time

    # use current time as a unique identifier for every folder created
    current_time = datetime.datetime.now()
    folder_name = f'RESULTS_{time.time()}'

    desktop_path = Path.home() / "Desktop"  # Get desktop path
    folder_path = desktop_path / folder_name  # Full path to new folder
    
    folder_path.mkdir(parents=True, exist_ok=True) # create folder
    print(f"Folder created at: {folder_path}")

    return folder_path



def groupby_thickness(new_data, best_model_list, thickness_list, temp):
    '''
    '''
    for i in range(len(thickness_list)):
            if thickness_list[i] not in new_data.keys():
                new_data[thickness_list[i]] = {
                    temp: best_model_list[i]
                }
            else:
                new_data[thickness_list[i]][temp] = best_model_list[i]

    return 0


def plots_data_writer(data, file_path, sheet_name, headers):
    """
    Writes data to a new sheet in an Excel file.
    Creates the file if it doesn't exist.
        
    Args:
        data (dict): Data to write (rows of columns)
        file_path (str): Path to the Excel file
        sheet_name (str): name for the sheet
        headers (list): dynammic headers for the sheet
    """

    # check if file exists
    if os.path.exists(file_path):
        workbook = load_workbook(file_path)
    else:
        workbook = Workbook()

        # Remove default sheet if it's empty
        if 'Sheet' in workbook.sheetnames:
            workbook.remove(workbook['Sheet'])

    # Add new sheet
    ws = workbook.create_sheet(title=sheet_name)
    
    # Styles
    bold_font = Font(bold=True)
    thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
    )
     
    center_align = Alignment(horizontal='center', vertical='center')

    # Extract data
    thickness = list(data.keys())[0]
    temperatures = list(data[thickness].keys())
    num_temps = len(temperatures)
    total_columns = num_temps * 3

    # Row 1: Thickness
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_columns)
    cell = ws.cell(row=1, column=1)
    cell.value = f"{thickness} mm"
    cell.font = bold_font
    cell.alignment = center_align
    cell.border = thin_border

    # Row 2-3: Temperature + Subheaders
    for i, temp in enumerate(temperatures):
        col_base = i * 3 + 1

        # Temperature header
        ws.merge_cells(start_row=2, start_column=col_base, end_row=2, end_column=col_base+2)
        temp_cell = ws.cell(row=2, column=col_base)
        temp_cell.value = f"{temp} Degrees Celcius"
        temp_cell.font = bold_font
        temp_cell.alignment = center_align
        temp_cell.border = thin_border

        for col in range(col_base, col_base + 3):
            ws.cell(row=2, column=col).border = thin_border

         # Subheaders
        headers = headers
        for j, h in enumerate(headers):
            sub_cell = ws.cell(row=3, column=col_base + j)
            sub_cell.value = h
            sub_cell.font = bold_font
            sub_cell.alignment = center_align
            sub_cell.border = thin_border

        # Data rows
        data_rows = list(data[thickness][temp].keys())
        time = data[thickness][temp][data_rows[0]]
        data1 = data[thickness][temp][data_rows[1]]
        data2 = data[thickness][temp][data_rows[2]]

        for r_idx in range(len(time)):
            r = r_idx + 4
            ws.cell(row=r, column=col_base).value = time[r_idx]
            ws.cell(row=r, column=col_base+1).value = data1[r_idx]
            ws.cell(row=r, column=col_base+2).value = data2[r_idx]

            for j in range(3):
                d_cell = ws.cell(row=r, column=col_base + j)
                d_cell.border = thin_border
                d_cell.alignment = center_align

    # Apply borders to thickness and headers
    for row in range(1, 4):
        for col in range(1, total_columns + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
            cell.alignment = center_align

    # Auto-resize columns
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_length + 2
        
    # save the workbook
    workbook.save(file_path)
    return

def plot_drying_curve(best_model_results, folder_path, file_path):
    '''
    This function plots and saves the drying curve
    to folder path using the data in best_model_results
    '''

    # colors = ['red', 'black', 'blue', 'green', 'purple']
    linestyle = ['--', '-.', '-', ':']
    markers = ['v', '*', '+', 'x', '^','o']

    
    for thickness in best_model_results.keys():
        data = {
            thickness: {}
        }

        plt.figure(figsize=(10, 6))

        for temp in best_model_results[thickness].keys():
            
            # extract time and moisture ratio (experimental and predicted)
            time = best_model_results[thickness][temp]['time']
            MR1 = best_model_results[thickness][temp]['MR1'] # experimental moisture ratio
            MR2 = best_model_results[thickness][temp]['MR2'] # predicted moisture ratio
 
            MR_ = []
            for i in range(0, max(time)+1):
                if i in time:
                    MR_.append(MR1[np.where(time == i)[0][0]])
                else:
                    MR_.append('')

            data[thickness][temp] = {
                'time': list(range(0, max(time) + 1)),
                'MR_ex': MR_,
                'MR_pre': MR2
            }

  
            
            random.shuffle(markers)
            plt.scatter(time, MR1, label=f'Experimental: {temp} Degrees Celcius', 
                marker=random.choice(markers))
                       

            plt.plot(list(range(0, max(time)+1)), MR2, label=f'Predicted: {temp} Degrees Celcius',
                linestyle=random.choice(linestyle))


        plots_data_writer(data, file_path,
            f'Drying curve ({thickness}mm)',
            ['Time (mins)', 'MR (Experimental)', 'MR (Predicted)']
        )

        plt.legend(fontsize=10, title_fontsize=12, loc="upper right")

        plt.xlabel("Time (mins)", fontsize=12)
        plt.ylabel("Moisture Ratio (Dry Basis)", fontsize=12)
        plt.title(f'Drying Curve for {thickness}mm thickness')

        # Optional: grid for better readability
        plt.grid(True, linestyle='--', alpha=0.5)

        # Display the plot
        plt.tight_layout()
        plt.savefig(f'{folder_path}/Drying_Curve_{thickness}.jpg', dpi=300)
        plt.close()


def plot_drying_rate_curve(best_model_results, folder_path, file_path):
    '''
    '''

    for thickness in best_model_results.keys():
        data = {
            thickness: {}
        }

        plt.figure(figsize=(10, 6))
        
        for temp in best_model_results[thickness].keys():
            
            # extract time and moisture ratio (predicted)
            time = best_model_results[thickness][temp]['time']

            time2 = np.array(range(0, max(time) + 1))
            MR2 = best_model_results[thickness][temp]['MR2'] # predicted moisture ratio
            
            drying_rate = -np.diff(MR2) / np.diff(time2)
            drying_rate_arr = np.insert(drying_rate, 0, 0) # insert 0 at index 0

            data[thickness][temp] = {
                'time': time2,
                'MR_pre': MR2,
                'Drying Rate': drying_rate_arr
            
            }


            plt.plot(time2, drying_rate_arr, label=f'{temp} Degrees Celcius')

        plots_data_writer(data, file_path,
            f'Drying rate curve ({thickness}mm)',
            ['Time (mins)', 'MR (Predicted)', 'Drying Rate']
        )

        plt.legend(fontsize=10, title_fontsize=12, loc="upper right")

        plt.xlabel("Time (mins)", fontsize=12)
        plt.ylabel("Drying Rate (per mins)", fontsize=12)
        plt.title(f'Drying Rate Curve ({thickness}mm)')

        # Optional: grid for better readability
        plt.grid(True, linestyle='--', alpha=0.5)

        # Display the plot
        plt.tight_layout()
        plt.savefig(f'{folder_path}/Drying_Rate_Curve_{thickness}.jpg', dpi=300)
        plt.close()


def  plot_krischer_curve(best_model_results, folder_path):
    '''
    '''

    for thickness in best_model_results.keys():
        data = {}
        
        plt.figure(figsize=(10, 6))

        for temp in best_model_results[thickness].keys():
        
            # extract time and moisture ratio (predicted)
            time = best_model_results[thickness][temp]['time']

            time2 = np.array(range(0, max(time) + 1))
            MR2 = best_model_results[thickness][temp]['MR2'] # predicted moisture ratio
            
            drying_rate = -np.diff(MR2) / np.diff(time2)
            drying_rate_arr = np.insert(drying_rate, 0, 0) # insert 0 at index 0

            
            plt.plot(MR2, drying_rate_arr, label=f'{temp} Degrees Celcius')


        plt.legend(fontsize=10, title_fontsize=12, loc="upper left")

        plt.xlabel("Moisture ratio (Dry Basis)", fontsize=12)
        plt.ylabel("Drying Rate (per mins)", fontsize=12)
        plt.title(f'Krischer Curve ({thickness}mm)')

        # Optional: grid for better readability
        plt.grid(True, linestyle='--', alpha=0.5)

        # Display the plot
        plt.tight_layout()
        plt.savefig(f'{folder_path}/Krischer_Curve_{thickness}.jpg', dpi=300)
        plt.close()

    return 0



def plot_handler(best_model_results, folder_path, file_path):
    '''
    This function handles the drawing of the 
    required plots: Drying curve,
    Drying rate curve and Krischer curve.
    '''
    
    # plot drying curve
    plot_drying_curve(best_model_results, folder_path, file_path)

    # plot drying rate curve
    plot_drying_rate_curve(best_model_results, folder_path, file_path)
    
    # plot krischer curve
    plot_krischer_curve(best_model_results, folder_path)

    
    return 0
